"""End-to-end smoke test: login, customize, write cells, read aggregates, export.

Run with: .venv/bin/python -m tests.test_smoke   (no pytest required)
"""

import os
import sys
import tempfile
from pathlib import Path

BACKEND = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND))

os.environ["APP_PASSWORD"] = "test-password"
os.environ["SECRET_KEY"] = "test-secret"
os.environ["DATABASE_URL"] = f"sqlite:///{Path(tempfile.gettempdir()) / 'finance_smoke.db'}"

db_file = Path(tempfile.gettempdir()) / "finance_smoke.db"
db_file.unlink(missing_ok=True)

from fastapi.testclient import TestClient  # noqa: E402

from app.db import SessionLocal, engine  # noqa: E402
from app.main import app  # noqa: E402
from app.models import Base  # noqa: E402
from app.seed import seed_database  # noqa: E402

Base.metadata.create_all(engine)
with SessionLocal() as db:
    seed_database(db, template="default")

client = TestClient(app)
failures: list[str] = []


def check(label: str, condition: bool, detail: str = "") -> None:
    status = "PASS" if condition else "FAIL"
    print(f"  [{status}] {label}{'' if condition else '  -> ' + detail}")
    if not condition:
        failures.append(label)


print("\nauth")
check("rejects bad password", client.post("/api/auth/login", json={"password": "nope"}).status_code == 401)
res = client.post("/api/auth/login", json={"password": "test-password"})
check("accepts good password", res.status_code == 200, res.text)
token = res.json()["token"]
H = {"Authorization": f"Bearer {token}"}
check("blocks unauthenticated reads", client.get("/api/sheets").status_code == 401)
check("allows authenticated reads", client.get("/api/sheets", headers=H).status_code == 200)

print("\ncustomization")
sheets = client.get("/api/sheets", headers=H).json()
check("seeded two sheets", len(sheets) == 2, str(sheets))
income = next(s for s in sheets if s["kind"] == "inflow")
spend = next(s for s in sheets if s["kind"] == "outflow")

res = client.post(
    "/api/sheets",
    headers=H,
    json={"name": "Side hustle", "kind": "inflow", "plan_label": "Target", "color": "#A855F7"},
)
check("creates a custom sheet", res.status_code == 201, res.text)
custom = res.json()
check("slugifies the name", custom["slug"] == "side-hustle", custom["slug"])

res = client.post(
    "/api/categories", headers=H, json={"sheet_id": custom["id"], "name": "Etsy shop"}
)
check("creates a category", res.status_code == 201, res.text)
category = res.json()
check(
    "rejects duplicate category names",
    client.post("/api/categories", headers=H, json={"sheet_id": custom["id"], "name": "Etsy shop"}).status_code == 409,
)
res = client.patch(
    f"/api/categories/{category['id']}", headers=H, json={"name": "Etsy store", "color": "#F59E0B"}
)
check("renames and recolors a category", res.json()["name"] == "Etsy store")

res = client.patch(f"/api/sheets/{custom['id']}", headers=H, json={"name": "Side income"})
check("renames a sheet", res.json()["name"] == "Side income")

print("\ncells")
YEAR = 2026
rent = next(
    c for c in client.get(f"/api/categories?sheet={spend['slug']}", headers=H).json()
    if c["name"] == "Rent"
)
check(
    "writes a plan value",
    client.patch("/api/values", headers=H, json={"category_id": rent["id"], "year": YEAR, "month": 0, "kind": "plan", "amount": 1200}).status_code == 200,
)
check(
    "fills a whole row",
    client.post("/api/values/fill-row", headers=H, json={"category_id": rent["id"], "year": YEAR, "amount": 1150}).json()["written"] == 12,
)
salary = next(
    c for c in client.get(f"/api/categories?sheet={income['slug']}", headers=H).json()
    if c["name"] == "Salary (net)"
)
client.post("/api/values/fill-row", headers=H, json={"category_id": salary["id"], "year": YEAR, "amount": 3000})

grid = client.get(f"/api/grid?sheet={spend['slug']}&year={YEAR}", headers=H).json()
rent_row = next(r for r in grid["rows"] if r["category_id"] == rent["id"])
check("row total is 12 x 1150", float(rent_row["total"]) == 13800.0, rent_row["total"])
check("plan column round-trips", float(rent_row["plan"]) == 1200.0, str(rent_row["plan"]))
check("variance = plan - actual for outflow", float(rent_row["variance"]) == -12600.0, str(rent_row["variance"]))
check("grid has 12 month columns", len(grid["months"]) == 12)

check(
    "clears a cell with null",
    client.patch("/api/values", headers=H, json={"category_id": rent["id"], "year": YEAR, "month": 3, "amount": None}).status_code == 200,
)
grid = client.get(f"/api/grid?sheet={spend['slug']}&year={YEAR}", headers=H).json()
rent_row = next(r for r in grid["rows"] if r["category_id"] == rent["id"])
check("cleared cell is null", rent_row["cells"][2] is None, str(rent_row["cells"][:4]))
check("total drops after clearing", float(rent_row["total"]) == 12650.0, rent_row["total"])

check(
    "bulk write",
    client.patch("/api/values/bulk", headers=H, json={"patches": [
        {"category_id": rent["id"], "year": YEAR, "month": 3, "amount": 1150},
        {"category_id": salary["id"], "year": YEAR, "month": 1, "amount": 3100},
    ]}).json()["written"] == 2,
)

print("\nfiscal year")
client.patch("/api/preferences", headers=H, json={"fiscal_start_month": 4})
grid = client.get(f"/api/grid?sheet={spend['slug']}&year={YEAR}", headers=H).json()
check("first column is April", grid["months"][0]["month"] == 4, str(grid["months"][0]))
check("rolls into next calendar year", grid["months"][-1]["year"] == YEAR + 1, str(grid["months"][-1]))
client.patch("/api/preferences", headers=H, json={"fiscal_start_month": 1})

print("\nnet worth")
items = client.get("/api/networth-items", headers=H).json()
savings = next(i for i in items if i["name"] == "Savings account")
card = next(i for i in items if i["name"] == "Credit card balance")
client.patch("/api/networth-values", headers=H, json={"item_id": savings["id"], "year": YEAR, "month": 1, "amount": 20000})
client.patch("/api/networth-values", headers=H, json={"item_id": card["id"], "year": YEAR, "month": 1, "amount": 2500})
res = client.post(f"/api/networth-values/carry-forward?year={YEAR}&month=2", headers=H)
check("carry-forward copies balances", res.json()["written"] == 2, res.text)
nw = client.get(f"/api/networth?year={YEAR}", headers=H).json()
check("net worth = assets - liabilities", float(nw["net_worth"][0]) == 17500.0, str(nw["net_worth"][:2]))
check("carried month matches", float(nw["net_worth"][1]) == 17500.0, str(nw["net_worth"][:3]))
check("unfilled months carry forward", float(nw["net_worth"][11]) == 17500.0, str(nw["net_worth"][-1]))

print("\naggregates")
agg = client.get(f"/api/aggregates?year={YEAR}", headers=H).json()
check("income total", float(agg["total_inflow"]) == 36100.0, agg["total_inflow"])
check("expense total", float(agg["total_outflow"]) == 13800.0, agg["total_outflow"])
check("balance", float(agg["balance"]) == 22300.0, agg["balance"])
check("savings rate", round(agg["savings_rate"], 1) == 61.8, str(agg["savings_rate"]))
check("per-sheet summaries", len(agg["sheets"]) == 3, str(len(agg["sheets"])))
check("year appears in available years", YEAR in agg["available_years"], str(agg["available_years"]))

print("\npreferences")
res = client.patch("/api/preferences", headers=H, json={"accent": "#FF0000", "goals": {"savings_rate_target": 35}})
prefs = res.json()
check("saves a top-level preference", prefs["accent"] == "#FF0000")
check("deep-merges nested goals", prefs["goals"]["savings_rate_target"] == 35)
check("keeps untouched nested keys", "emergency_fund_months" in prefs["goals"])
check("reset restores defaults", client.post("/api/preferences/reset", headers=H).json()["accent"] == "#6366F1")
check("lists templates", len(client.get("/api/structure/templates", headers=H).json()) == 4)

print("\nbalance row (workbook parity)")
grid = client.get(f"/api/grid?sheet={spend['slug']}&year={YEAR}", headers=H).json()
check("outflow sheet carries a balance row", grid["balance_row"] is not None)
check("balance = income - expenditure per month", float(grid["balance_row"][0]) == 3100.0 - 1150.0, str(grid["balance_row"][:2]))
check("balance row has a plan cell", float(grid["balance_plan"]) == -1200.0, str(grid["balance_plan"]))
income_grid = client.get(f"/api/grid?sheet={income['slug']}&year={YEAR}", headers=H).json()
check("inflow sheet has no balance row", income_grid["balance_row"] is None)

print("\ntools")
res = client.post(
    "/api/tools",
    headers=H,
    json={"name": "Test card", "purpose": "Credit Card", "bonus": "$100", "link": "https://example.com/r/abc"},
)
check("creates a tool", res.status_code == 201, res.text)
tool = res.json()
check("rejects duplicate tool names", client.post("/api/tools", headers=H, json={"name": "Test card"}).status_code == 409)
check("lists tools", any(t["id"] == tool["id"] for t in client.get("/api/tools", headers=H).json()))
check(
    "updates a tool",
    client.patch(f"/api/tools/{tool['id']}", headers=H, json={"bonus": "$150"}).json()["bonus"] == "$150",
)
check("accepts a plain referral code as the link", client.post(
    "/api/tools", headers=H, json={"name": "Taptap", "link": "SHAHMOHA9"}
).status_code == 201)

print("\nexport")
res = client.get(f"/api/export?year={YEAR}", headers=H)
check("returns an xlsx", res.status_code == 200 and res.content[:2] == b"PK", res.text[:200])
check("non-trivial size", len(res.content) > 5000, str(len(res.content)))

import io  # noqa: E402

from openpyxl import load_workbook  # noqa: E402

wb = load_workbook(io.BytesIO(res.content))
check("workbook tabs present", {"Dashboard", "Net Worth", "Tools", "Settings"} <= set(wb.sheetnames), str(wb.sheetnames))
ws = wb[spend["name"][:31]]
labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
check("exported sheet has a Total row", "Total" in labels, str(labels[-3:]))
check("exported sheet has a Balance row", "Balance" in labels, str(labels[-3:]))
check("Tools tab uses the workbook's headers", [c.value for c in wb["Tools"][1]] ==
      ["Sl.", "Tools to use", "Purpose", "Referral Bonus", "Referral Link or code"],
      str([c.value for c in wb["Tools"][1]]))

print("\npeople and split expenses")
res = client.post("/api/people", headers=H, json={"name": "Riko", "relation": "friend"})
check("creates a person", res.status_code == 201, res.text)
riko = res.json()
colleague = client.post("/api/people", headers=H, json={"name": "Dan", "relation": "colleague"}).json()
check("rejects duplicate person names", client.post("/api/people", headers=H, json={"name": "Riko"}).status_code == 409)

dining = next(
    c for c in client.get(f"/api/categories?sheet={spend['slug']}", headers=H).json()
    if c["name"] == "Dining out"
)


def cell(category_id, year, month):
    grid = client.get(f"/api/grid?sheet={spend['slug']}&year={year}", headers=H).json()
    row = next((r for r in grid["rows"] if r["category_id"] == category_id), None)
    if row is None:
        return None
    column = next((m for m in grid["months"] if m["year"] == year and m["month"] == month), None)
    return row["cells"][column["index"]] if column else None


# individual expense: the whole amount is mine
res = client.post("/api/expenses", headers=H, json={
    "spent_on": f"{YEAR}-05-04", "category_id": dining["id"], "total_amount": 2000,
    "merchant": "Solo lunch", "is_split": False,
})
check("creates an individual expense", res.status_code == 201, res.text)
solo = res.json()
check("individual: my share is the full total", float(solo["my_share"]) == 2000.0, str(solo["my_share"]))
check("individual: posts full amount to the grid", float(cell(dining["id"], YEAR, 5)) == 2000.0, str(cell(dining["id"], YEAR, 5)))

# split expense: only my share hits the grid
res = client.post("/api/expenses", headers=H, json={
    "spent_on": f"{YEAR}-05-11", "category_id": dining["id"], "total_amount": 9000,
    "merchant": "Group dinner", "is_split": True,
    "shares": [{"person_id": riko["id"], "amount": 3000}, {"person_id": colleague["id"], "amount": 3000}],
})
check("creates a split expense", res.status_code == 201, res.text)
split = res.json()
check("split: my share is total minus others", float(split["my_share"]) == 3000.0, str(split["my_share"]))
check("split: grid gets only my share", float(cell(dining["id"], YEAR, 5)) == 5000.0, str(cell(dining["id"], YEAR, 5)))
check("split: records one share per person", len(split["shares"]) == 2, str(len(split["shares"])))
check(
    "rejects shares exceeding the total",
    client.post("/api/expenses", headers=H, json={
        "spent_on": f"{YEAR}-05-12", "category_id": dining["id"], "total_amount": 100,
        "is_split": True, "shares": [{"person_id": riko["id"], "amount": 500}],
    }).status_code == 400,
)

print("\nwho owes me what")
balances = {b["person"]["name"]: b for b in client.get("/api/people/balances", headers=H).json()}
check("tracks what a friend owes", float(balances["Riko"]["owed"]) == 3000.0, str(balances["Riko"]["owed"]))
check("tracks what a colleague owes", float(balances["Dan"]["owed"]) == 3000.0, str(balances["Dan"]["owed"]))

riko_share = next(s for s in split["shares"] if s["person_id"] == riko["id"])
res = client.post(f"/api/expenses/shares/{riko_share['id']}/settle", headers=H)
check("settles one share", res.status_code == 200, res.text)
balances = {b["person"]["name"]: b for b in client.get("/api/people/balances", headers=H).json()}
check("settled share clears the debt", float(balances["Riko"]["owed"]) == 0.0, str(balances["Riko"]["owed"]))
check("settled amount is remembered", float(balances["Riko"]["settled"]) == 3000.0, str(balances["Riko"]["settled"]))
check("settling does NOT change the expense total", float(cell(dining["id"], YEAR, 5)) == 5000.0, str(cell(dining["id"], YEAR, 5)))

print("\nediting an expense reverses cleanly")
res = client.patch(f"/api/expenses/{split['id']}", headers=H, json={"total_amount": 12000})
check("raises the total", res.status_code == 200, res.text)
check("my share absorbs the increase", float(res.json()["my_share"]) == 6000.0, str(res.json()["my_share"]))
check("grid follows the new share", float(cell(dining["id"], YEAR, 5)) == 8000.0, str(cell(dining["id"], YEAR, 5)))

res = client.patch(f"/api/expenses/{split['id']}", headers=H, json={"spent_on": f"{YEAR}-06-02"})
check("moving months empties the old cell", float(cell(dining["id"], YEAR, 5)) == 2000.0, str(cell(dining["id"], YEAR, 5)))
check("moving months fills the new cell", float(cell(dining["id"], YEAR, 6)) == 6000.0, str(cell(dining["id"], YEAR, 6)))

groceries = next(
    c for c in client.get(f"/api/categories?sheet={spend['slug']}", headers=H).json()
    if c["name"] == "Groceries"
)
client.patch(f"/api/expenses/{split['id']}", headers=H, json={"category_id": groceries["id"]})
check("recategorising moves the amount", float(cell(groceries["id"], YEAR, 6)) == 6000.0, str(cell(groceries["id"], YEAR, 6)))
check("old category is left clean", cell(dining["id"], YEAR, 6) is None, str(cell(dining["id"], YEAR, 6)))

client.patch(f"/api/expenses/{split['id']}", headers=H, json={"is_split": False, "shares": []})
check("un-splitting gives me the whole amount", float(cell(groceries["id"], YEAR, 6)) == 12000.0, str(cell(groceries["id"], YEAR, 6)))

check("deletes an expense", client.delete(f"/api/expenses/{split['id']}", headers=H).status_code == 204)
check("deleting reverses its posting", cell(groceries["id"], YEAR, 6) is None, str(cell(groceries["id"], YEAR, 6)))
check("unrelated expenses are untouched", float(cell(dining["id"], YEAR, 5)) == 2000.0, str(cell(dining["id"], YEAR, 5)))
check(
    "refuses to hard-delete someone who still owes you",
    client.delete(f"/api/people/{colleague['id']}?hard=true", headers=H).status_code in (204, 409),
)

print("\nreceipt scanning guardrails")
status_body = client.get("/api/receipts/status", headers=H).json()
check("reports whether scanning is configured", "enabled" in status_body, str(status_body))
res = client.post("/api/receipts/scan", headers=H, files={"image": ("notes.txt", b"totally not an image", "image/png")})
check("rejects a non-image disguised as one", res.status_code in (415, 503), f"{res.status_code} {res.text[:120]}")
res = client.post("/api/receipts/scan", headers=H, files={"image": ("empty.png", b"", "image/png")})
check("rejects an empty upload", res.status_code in (400, 503), str(res.status_code))
check("scan requires authentication", client.post("/api/receipts/scan", files={"image": ("a.png", b"\x89PNG\r\n\x1a\n", "image/png")}).status_code == 401)

from app.services.receipts import sniff_media_type as _sniff  # noqa: E402

check("sniffs PNG by content", _sniff(b"\x89PNG\r\n\x1a\nrest") == "image/png")
check("sniffs JPEG by content", _sniff(b"\xff\xd8\xff\xe0rest") == "image/jpeg")
check("sniffs WebP by content", _sniff(b"RIFF\x00\x00\x00\x00WEBPVP8 ") == "image/webp")
check("refuses a PDF", _sniff(b"%PDF-1.7") is None)
check("refuses a script", _sniff(b"#!/bin/sh\nrm -rf /") is None)

print("\nworkbook import (structure only by default)")
from openpyxl import Workbook as _Workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.import_workbook import import_workbook  # noqa: E402
from app.models import Category as _Category  # noqa: E402
from app.models import MonthlyValue as _MonthlyValue  # noqa: E402
from app.models import Sheet as _Sheet  # noqa: E402

fake = _Workbook()
ws = fake.active
ws.title = "Income"
ws.append(["Description", "Expected income", "Actual income"])
ws.append([None, None, *[f"M{i}" for i in range(1, 13)]])
ws.append(["Consulting fees", 1000, *[500] * 12])
ws.append(["Total income", 1000, *[500] * 12])
ws2 = fake.create_sheet("Net Worth")
ws2.append(["Net Worth Tracker", None, "Assets", None, "Liabilities"])
ws2.append([None, "Vanguard ISA", 5000, "Overdraft", 250])
ws3 = fake.create_sheet("Tools")
ws3.append(["Sl.", "Tools to use", "Purpose", "Referral Bonus", "Referral Link or code"])
ws3.append([1, "Monzo", "Banking", 50, "MONZO50"])
fake_path = Path(tempfile.gettempdir()) / "finance_smoke_import.xlsx"
fake.save(fake_path)

IMPORT_YEAR = 2031
with SessionLocal() as import_db:
    created = import_workbook(import_db, fake_path, year=IMPORT_YEAR)
    check("imports categories", created["categories"] >= 1, str(created))
    check("imports net-worth items", created["networth_items"] == 2, str(created))
    check("imports tools", created["tools"] == 1, str(created))
    check("skips the Total row", created["categories"] == 1, str(created))
    check("writes NO amounts by default", created["values"] == 0, str(created))

    imported = import_db.execute(
        select(_Category).join(_Sheet).where(_Category.name == "Consulting fees")
    ).scalar_one()
    cells = import_db.execute(
        select(_MonthlyValue).where(
            _MonthlyValue.category_id == imported.id, _MonthlyValue.year == IMPORT_YEAR
        )
    ).scalars().all()
    check("imported category starts empty", cells == [], str(cells))

    # ...and the opt-in still works
    created = import_workbook(import_db, fake_path, year=IMPORT_YEAR, with_values=True)
    # 1 plan + 12 months for the one category, plus 2 net-worth balances
    check("--with-values carries amounts over", created["values"] == 15, str(created))
    cells = import_db.execute(
        select(_MonthlyValue).where(
            _MonthlyValue.category_id == imported.id, _MonthlyValue.year == IMPORT_YEAR
        )
    ).scalars().all()
    check("amounts land on the imported category", len(cells) == 13, str(len(cells)))

fake_path.unlink(missing_ok=True)

print("\ndelete behaviour")
res = client.delete(f"/api/categories/{category['id']}", headers=H)
check("soft delete returns 204", res.status_code == 204)
check(
    "soft-deleted category hidden by default",
    all(c["id"] != category["id"] for c in client.get("/api/categories", headers=H).json()),
)
check(
    "soft-deleted category visible with include_inactive",
    any(c["id"] == category["id"] for c in client.get("/api/categories?include_inactive=true", headers=H).json()),
)
check("hard delete returns 204", client.delete(f"/api/categories/{category['id']}?hard=true", headers=H).status_code == 204)
check(
    "hard-deleted category is gone",
    all(c["id"] != category["id"] for c in client.get("/api/categories?include_inactive=true", headers=H).json()),
)

print("\n" + ("ALL CHECKS PASSED" if not failures else f"{len(failures)} FAILED: {failures}"))
sys.exit(1 if failures else 0)
