"""Excel export — the backup ritual. One .xlsx with every sheet, net worth,
a dashboard summary, and the settings used to build it."""

from typing import Optional

import io
import logging
from datetime import datetime
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..auth import require_auth
from ..db import get_db
from ..preferences import get_preferences
from .aggregates import read_aggregates
from .grid import read_grid
from .networth import build_networth
from .structure import list_sheets, list_tools

log = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["export"], dependencies=[Depends(require_auth)])

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(bottom=THIN)


def _number_format(prefs: dict) -> str:
    decimals = int(prefs.get("decimals", 0) or 0)
    symbol = prefs.get("currency_symbol", "")
    body = "#,##0" + ("." + "0" * decimals if decimals else "")
    if not symbol:
        return body
    return f'"{symbol}"{body}' if prefs.get("symbol_position") == "before" else f'{body}"{symbol}"'


def _style_header(ws, row: int, width: int) -> None:
    for col in range(1, width + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, first_width: int = 34, rest: int = 13) -> None:
    ws.column_dimensions["A"].width = first_width
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = rest
    ws.freeze_panes = "B2"


def _f(value: Optional[Decimal]) -> Optional[float]:
    return float(value) if value is not None else None


@router.get("/export")
def export_workbook(
    year: int = Query(..., ge=1900, le=2999),
    include_inactive: bool = False,
    db: Session = Depends(get_db),
):
    prefs = get_preferences(db)
    fmt = _number_format(prefs)
    wb = Workbook()
    wb.remove(wb.active)

    aggregates = read_aggregates(year=year, db=db)
    month_labels = [m.label for m in aggregates.months]

    # ---------------------------------------------------------- Dashboard
    ws = wb.create_sheet("Dashboard")
    ws.append([f"{prefs.get('app_name', 'Finance')} — {year}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Total income", _f(aggregates.total_inflow)])
    ws.append(["Total expenses", _f(aggregates.total_outflow)])
    ws.append(["Balance", _f(aggregates.balance)])
    ws.append(["Savings rate (%)", aggregates.savings_rate])
    ws.append(["Net worth", _f(aggregates.net_worth_latest)])
    ws.append(["Assets", _f(aggregates.assets_latest)])
    ws.append(["Liabilities", _f(aggregates.liabilities_latest)])
    for row in range(3, 10):
        ws.cell(row=row, column=1).font = TOTAL_FONT
        if row != 6:
            ws.cell(row=row, column=2).number_format = fmt

    ws.append([])
    start = ws.max_row + 1
    ws.append(["Monthly", *month_labels, "Total"])
    _style_header(ws, start, len(month_labels) + 2)
    for label, series, total in (
        ("Income", aggregates.inflow_monthly, aggregates.total_inflow),
        ("Expenses", aggregates.outflow_monthly, aggregates.total_outflow),
        ("Balance", aggregates.balance_monthly, aggregates.balance),
        ("Net worth", aggregates.net_worth_series, aggregates.net_worth_latest),
    ):
        ws.append([label, *[_f(v) for v in series], _f(total)])
        for col in range(2, len(month_labels) + 3):
            ws.cell(row=ws.max_row, column=col).number_format = fmt
    _autosize(ws)

    # ---------------------------------------------------------- one tab per sheet
    for sheet in list_sheets(include_inactive=include_inactive, db=db):
        grid = read_grid(
            sheet=sheet.slug, year=year, include_inactive=include_inactive, db=db
        )
        ws = wb.create_sheet(sheet.name[:31])
        header = ["Category", "Group"]
        if sheet.show_plan:
            header.append(sheet.plan_label)
        header += month_labels + ["Total", "Average"]
        if sheet.show_plan:
            header.append("Variance")
        ws.append(header)
        _style_header(ws, 1, len(header))

        for row in grid.rows:
            line: list = [row.name, row.group_name or ""]
            if sheet.show_plan:
                line.append(_f(row.plan))
            line += [_f(c) for c in row.cells]
            line += [_f(row.total), _f(row.average)]
            if sheet.show_plan:
                line.append(_f(row.variance))
            ws.append(line)
            for col in range(3, len(header) + 1):
                ws.cell(row=ws.max_row, column=col).number_format = fmt

        totals: list = ["Total", ""]
        if sheet.show_plan:
            totals.append(_f(grid.plan_total))
        totals += [_f(c) for c in grid.column_totals]
        totals += [_f(grid.grand_total), None]
        if sheet.show_plan:
            totals.append(_f(grid.plan_total - grid.grand_total))
        ws.append(totals)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = TOTAL_FONT
            cell.border = BORDER
            if col >= 3:
                cell.number_format = fmt

        # The workbook's Balance row, on the expenditure-side tabs.
        if grid.balance_row is not None:
            balance: list = ["Balance", ""]
            if sheet.show_plan:
                balance.append(_f(grid.balance_plan))
            balance += [_f(v) for v in grid.balance_row]
            balance.append(_f(sum(grid.balance_row)))
            ws.append(balance)
            for col in range(1, len(header) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = TOTAL_FONT
                if col >= 3:
                    cell.number_format = fmt
        _autosize(ws)

    # ---------------------------------------------------------- Net worth
    networth = build_networth(db, year, include_inactive)
    ws = wb.create_sheet("Net Worth")
    ws.append(["Item", *month_labels, "Latest", "Change"])
    _style_header(ws, 1, len(month_labels) + 3)

    def _block(title: str, rows, totals) -> None:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = TOTAL_FONT
        for row in rows:
            ws.append([row.name, *[_f(c) for c in row.cells], _f(row.latest), _f(row.change)])
            for col in range(2, len(month_labels) + 4):
                ws.cell(row=ws.max_row, column=col).number_format = fmt
        ws.append([f"Total {title.lower()}", *[_f(t) for t in totals]])
        for col in range(1, len(month_labels) + 2):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = TOTAL_FONT
            cell.border = BORDER
            if col > 1:
                cell.number_format = fmt
        ws.append([])

    _block("Assets", networth.assets, networth.asset_totals)
    _block("Liabilities", networth.liabilities, networth.liability_totals)
    ws.append(["Net worth", *[_f(v) for v in networth.net_worth]])
    for col in range(1, len(month_labels) + 2):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = TOTAL_FONT
        if col > 1:
            cell.number_format = fmt
    _autosize(ws)

    # ---------------------------------------------------------- Tools
    tools = list_tools(include_inactive=include_inactive, db=db)
    if tools:
        ws = wb.create_sheet("Tools")
        ws.append(["Sl.", "Tools to use", "Purpose", "Referral Bonus", "Referral Link or code"])
        _style_header(ws, 1, 5)
        for index, tool in enumerate(tools, start=1):
            ws.append([index, tool.name, tool.purpose or "", tool.bonus or "", tool.link or ""])
        ws.column_dimensions["A"].width = 6
        for column, width in (("B", 32), ("C", 22), ("D", 18), ("E", 70)):
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"

    # ---------------------------------------------------------- Settings snapshot
    ws = wb.create_sheet("Settings")
    ws.append(["Key", "Value"])
    _style_header(ws, 1, 2)
    for key, value in sorted(prefs.items()):
        ws.append([key, str(value)])
    ws.append(["exported_at", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws, first_width=28, rest=48)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{prefs.get('app_name', 'finance')}-{year}.xlsx".replace(" ", "-").lower()
    log.info("export generated year=%s file=%s", year, filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
