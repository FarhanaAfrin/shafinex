"""Seed from the original Excel workbook instead of hand-typing ~90 names.

Reads the "Personal Finance Tracker - Stat with SHAH" layout:

  Income / Expenditure   row 1 headers, row 2 month names in C..N,
                         data from row 3: A = name, B = plan, C..N = actuals,
                         trailing "Total …" / "Balance" rows ignored
  Net Worth              B = asset names (C = amount), D = liability names (E = amount)

Usage:
    python -m app.import_workbook "../Personal Finance Tracker - Stat with SHAH.xlsx"
    python -m app.import_workbook <path> --year 2026 --replace
    python -m app.import_workbook <path> --with-values      # also copy the amounts

**Structure only by default.** The import sets up the sheets, categories,
net-worth items and tools; the amounts are meant to be typed into the app, into
empty placeholder cells. Pass `--with-values` if you would rather carry the
workbook's numbers over as a starting point.
"""

import argparse
import logging
import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import Category, MonthlyValue, NetworthItem, NetworthValue, Sheet, Tool
from .seed import _wipe

log = logging.getLogger(__name__)

# Rows that are spreadsheet arithmetic, not categories.
SKIP = re.compile(r"^\s*(total|balance|net\s*worth|grand\s*total|sum)\b", re.IGNORECASE)

SHEET_SPECS = [
    {
        "tab": "Income",
        "slug": "income",
        "name": "Income",
        "kind": "inflow",
        "icon": "mdi-trending-up",
        "color": "#22C55E",
        "fallback_plan_label": "Expected",
    },
    {
        "tab": "Expenditure",
        "slug": "expenditure",
        "name": "Expenditure",
        "kind": "outflow",
        "icon": "mdi-trending-down",
        "color": "#EF4444",
        "fallback_plan_label": "Budget",
    },
]


def _clean(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _amount(value) -> Optional[Decimal]:
    if value is None or isinstance(value, str) and not value.strip():
        return None
    try:
        amount = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
    return amount if amount != 0 else None


def _bonus(value) -> Optional[str]:
    """The Referral Bonus column mixes numbers (100) with text ("$30-40")."""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return _clean(value)


def _plan_label(raw: Optional[str], fallback: str) -> str:
    """"Budgeted Expenses" -> "Budget", "Expected income" -> "Expected"."""
    if not raw:
        return fallback
    first = raw.split()[0].strip().title()
    return {"Budgeted": "Budget"}.get(first, first)


def import_workbook(
    db: Session,
    path: Path,
    year: int,
    replace: bool = False,
    with_values: bool = False,
) -> dict:
    """Import the workbook's structure. Amounts only come across with
    `with_values=True` — normally they are entered in the app."""
    wb = load_workbook(path, data_only=True)
    if replace:
        _wipe(db)

    created = {"sheets": 0, "categories": 0, "networth_items": 0, "tools": 0, "values": 0}

    for order, spec in enumerate(SHEET_SPECS):
        if spec["tab"] not in wb.sheetnames:
            log.warning("workbook has no '%s' tab — skipping", spec["tab"])
            continue
        ws = wb[spec["tab"]]

        sheet = db.execute(select(Sheet).where(Sheet.slug == spec["slug"])).scalar_one_or_none()
        if sheet is None:
            sheet = Sheet(
                slug=spec["slug"],
                name=spec["name"],
                kind=spec["kind"],
                icon=spec["icon"],
                color=spec["color"],
                plan_label=_plan_label(_clean(ws.cell(1, 2).value), spec["fallback_plan_label"]),
                sort_order=order,
            )
            db.add(sheet)
            db.flush()
            created["sheets"] += 1

        existing = {
            name: cid
            for name, cid in db.execute(
                select(Category.name, Category.id).where(Category.sheet_id == sheet.id)
            )
        }
        position = len(existing)

        for row in range(3, ws.max_row + 1):
            name = _clean(ws.cell(row, 1).value)
            if not name or SKIP.match(name):
                continue

            category_id = existing.get(name)
            if category_id is None:
                category = Category(
                    sheet_id=sheet.id, name=name, sort_order=position, is_active=True
                )
                db.add(category)
                db.flush()
                category_id = category.id
                existing[name] = category_id
                position += 1
                created["categories"] += 1

            if not with_values:
                continue

            cells = [(0, "plan", _amount(ws.cell(row, 2).value))]
            cells += [
                (month, "actual", _amount(ws.cell(row, 2 + month).value))
                for month in range(1, 13)
            ]
            for month, kind, amount in cells:
                if amount is None:
                    continue
                found = db.execute(
                    select(MonthlyValue).where(
                        MonthlyValue.category_id == category_id,
                        MonthlyValue.year == year,
                        MonthlyValue.month == month,
                        MonthlyValue.kind == kind,
                    )
                ).scalar_one_or_none()
                if found is None:
                    db.add(
                        MonthlyValue(
                            category_id=category_id,
                            year=year,
                            month=month,
                            kind=kind,
                            amount=amount,
                        )
                    )
                else:
                    found.amount = amount
                created["values"] += 1

    # ------------------------------------------------------------ net worth
    if "Net Worth" in wb.sheetnames:
        ws = wb["Net Worth"]
        # (side, name column, amount column)
        for side, name_col, amount_col in (("asset", 2, 3), ("liability", 4, 5)):
            existing = {
                name: item_id
                for name, item_id in db.execute(
                    select(NetworthItem.name, NetworthItem.id).where(NetworthItem.side == side)
                )
            }
            position = len(existing)
            for row in range(2, ws.max_row + 1):
                name = _clean(ws.cell(row, name_col).value)
                if not name or SKIP.match(name):
                    continue

                item_id = existing.get(name)
                if item_id is None:
                    item = NetworthItem(side=side, name=name, sort_order=position)
                    db.add(item)
                    db.flush()
                    item_id = item.id
                    existing[name] = item_id
                    position += 1
                    created["networth_items"] += 1

                amount = _amount(ws.cell(row, amount_col).value) if with_values else None
                if amount is None:
                    continue
                # The workbook holds one snapshot, so it lands on the current month.
                month = date.today().month
                found = db.execute(
                    select(NetworthValue).where(
                        NetworthValue.item_id == item_id,
                        NetworthValue.year == year,
                        NetworthValue.month == month,
                    )
                ).scalar_one_or_none()
                if found is None:
                    db.add(
                        NetworthValue(item_id=item_id, year=year, month=month, amount=amount)
                    )
                else:
                    found.amount = amount
                created["values"] += 1

    # ------------------------------------------------------------ tools
    if "Tools" in wb.sheetnames:
        ws = wb["Tools"]
        existing = {name for name in db.execute(select(Tool.name)).scalars()}
        position = len(existing)
        for row in range(2, ws.max_row + 1):
            name = _clean(ws.cell(row, 2).value)
            if not name or name in existing:
                continue
            db.add(
                Tool(
                    name=name,
                    purpose=_clean(ws.cell(row, 3).value),
                    bonus=_bonus(ws.cell(row, 4).value),
                    link=_clean(ws.cell(row, 5).value),
                    sort_order=position,
                )
            )
            existing.add(name)
            position += 1
            created["tools"] = created.get("tools", 0) + 1

    db.commit()
    log.info("workbook import complete: %s", created)
    return created


def main() -> None:
    from .db import SessionLocal

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s %(message)s")
    parser = argparse.ArgumentParser(description="Import the original Excel workbook")
    parser.add_argument("path", type=Path, help="path to the .xlsx workbook")
    parser.add_argument("--year", type=int, default=date.today().year)
    parser.add_argument("--replace", action="store_true", help="wipe existing data first")
    parser.add_argument(
        "--with-values",
        action="store_true",
        help="also copy the workbook's amounts (default: structure only — you type the amounts in the app)",
    )
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"No such file: {args.path}")

    with SessionLocal() as db:
        created = import_workbook(
            db, args.path, year=args.year, replace=args.replace, with_values=args.with_values
        )
    print(
        f"Imported into {args.year}: {created['sheets']} sheets, "
        f"{created['categories']} categories, {created['networth_items']} net-worth items, "
        f"{created['tools']} tools, {created['values']} amounts"
    )
    if not args.with_values:
        print("Amounts left empty — enter them in the app (or re-run with --with-values).")


if __name__ == "__main__":
    main()
